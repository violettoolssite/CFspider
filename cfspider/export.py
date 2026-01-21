"""
CFspider 数据导出模块

支持导出数据到 JSON、CSV、Excel、SQLite 格式。

Example:
    >>> import cfspider
    >>> 
    >>> # 保存响应
    >>> response = cfspider.get("https://example.com")
    >>> response.save("page.html")
    >>> 
    >>> # 保存提取结果
    >>> data = response.pick(title="h1", price=".price")
    >>> data.save("output.csv")
    >>> 
    >>> # 使用导出函数
    >>> cfspider.export(data, "output.xlsx", format="excel")
"""

import json
import csv
import os
from typing import Any, Dict, List, Union, Optional


def export(data: Union[Dict, List[Dict], Any], 
           filepath: str, 
           format: str = None,
           table: str = "data",
           encoding: str = "utf-8",
           **kwargs) -> str:
    """
    导出数据到文件
    
    Args:
        data: 要导出的数据（字典、字典列表或其他）
        filepath: 输出文件路径
        format: 导出格式（json/csv/excel/sqlite），None 则自动从扩展名推断
        table: SQLite 表名（仅 sqlite 格式使用）
        encoding: 文件编码
        **kwargs: 传递给底层导出函数的参数
        
    Returns:
        输出文件的绝对路径
        
    Example:
        >>> export({"title": "Hello"}, "output.json")
        >>> export([{"a": 1}, {"a": 2}], "output.csv")
        >>> export(data, "output.xlsx", format="excel")
    """
    # 自动推断格式
    if format is None:
        ext = os.path.splitext(filepath)[1].lower()
        format_map = {
            '.json': 'json',
            '.csv': 'csv',
            '.xlsx': 'excel',
            '.xls': 'excel',
            '.db': 'sqlite',
            '.sqlite': 'sqlite',
            '.sqlite3': 'sqlite',
        }
        format = format_map.get(ext, 'json')
    
    format = format.lower()
    
    if format == 'json':
        return export_json(data, filepath, encoding=encoding, **kwargs)
    elif format == 'csv':
        return export_csv(data, filepath, encoding=encoding, **kwargs)
    elif format == 'excel':
        return export_excel(data, filepath, **kwargs)
    elif format == 'sqlite':
        return export_sqlite(data, filepath, table=table, **kwargs)
    else:
        raise ValueError(f"Unsupported format: {format}")


def export_json(data: Any, 
                filepath: str, 
                encoding: str = "utf-8",
                indent: int = 2,
                ensure_ascii: bool = False,
                **kwargs) -> str:
    """
    导出数据到 JSON 文件
    
    Args:
        data: 要导出的数据
        filepath: 输出文件路径
        encoding: 文件编码
        indent: 缩进空格数
        ensure_ascii: 是否转义非 ASCII 字符
        
    Returns:
        输出文件的绝对路径
    """
    filepath = os.path.abspath(filepath)
    
    with open(filepath, 'w', encoding=encoding) as f:
        json.dump(data, f, indent=indent, ensure_ascii=ensure_ascii, **kwargs)
    
    return filepath


def export_csv(data: Union[Dict, List[Dict]], 
               filepath: str,
               encoding: str = "utf-8-sig",  # 带 BOM 以支持 Excel 打开
               delimiter: str = ",",
               **kwargs) -> str:
    """
    导出数据到 CSV 文件
    
    Args:
        data: 要导出的数据（字典或字典列表）
        filepath: 输出文件路径
        encoding: 文件编码（默认 utf-8-sig 带 BOM）
        delimiter: 分隔符
        
    Returns:
        输出文件的绝对路径
    """
    filepath = os.path.abspath(filepath)
    
    # 统一转换为列表
    if isinstance(data, dict):
        # 单个字典，检查值是否为列表
        has_list_values = any(isinstance(v, list) for v in data.values())
        
        if has_list_values:
            # 展开列表值为多行
            rows = _expand_dict_with_lists(data)
        else:
            # 单行数据
            rows = [data]
    elif isinstance(data, list):
        rows = data
    else:
        rows = [{"value": data}]
    
    if not rows:
        return filepath
    
    # 获取所有字段名
    fieldnames = []
    for row in rows:
        if isinstance(row, dict):
            for key in row.keys():
                if key not in fieldnames:
                    fieldnames.append(key)
    
    with open(filepath, 'w', encoding=encoding, newline='') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, delimiter=delimiter, **kwargs)
        writer.writeheader()
        
        for row in rows:
            if isinstance(row, dict):
                # 将列表值转为字符串
                clean_row = {}
                for k, v in row.items():
                    if isinstance(v, (list, dict)):
                        clean_row[k] = json.dumps(v, ensure_ascii=False)
                    else:
                        clean_row[k] = v
                writer.writerow(clean_row)
    
    return filepath


def _expand_dict_with_lists(data: Dict) -> List[Dict]:
    """
    展开包含列表的字典为多行
    
    Example:
        >>> _expand_dict_with_lists({"title": "Hello", "links": ["a", "b", "c"]})
        [{"title": "Hello", "links": "a"}, {"title": "Hello", "links": "b"}, ...]
    """
    # 找出最长的列表长度
    max_len = 1
    for v in data.values():
        if isinstance(v, list):
            max_len = max(max_len, len(v))
    
    rows = []
    for i in range(max_len):
        row = {}
        for k, v in data.items():
            if isinstance(v, list):
                row[k] = v[i] if i < len(v) else None
            else:
                row[k] = v if i == 0 else None  # 非列表值只在第一行显示
        rows.append(row)
    
    return rows


def export_excel(data: Union[Dict, List[Dict]], 
                 filepath: str,
                 sheet_name: str = "Sheet1",
                 **kwargs) -> str:
    """
    导出数据到 Excel 文件
    
    Args:
        data: 要导出的数据（字典或字典列表）
        filepath: 输出文件路径
        sheet_name: 工作表名称
        
    Returns:
        输出文件的绝对路径
        
    Raises:
        ImportError: 如果未安装 openpyxl
    """
    try:
        from openpyxl import Workbook
    except ImportError:
        raise ImportError(
            "openpyxl is required for Excel export. "
            "Install it with: pip install openpyxl"
        )
    
    filepath = os.path.abspath(filepath)
    
    # 统一转换为列表
    if isinstance(data, dict):
        has_list_values = any(isinstance(v, list) for v in data.values())
        if has_list_values:
            rows = _expand_dict_with_lists(data)
        else:
            rows = [data]
    elif isinstance(data, list):
        rows = data
    else:
        rows = [{"value": data}]
    
    if not rows:
        wb = Workbook()
        wb.save(filepath)
        return filepath
    
    # 获取所有字段名
    fieldnames = []
    for row in rows:
        if isinstance(row, dict):
            for key in row.keys():
                if key not in fieldnames:
                    fieldnames.append(key)
    
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # 写入表头
    for col, name in enumerate(fieldnames, 1):
        ws.cell(row=1, column=col, value=name)
    
    # 写入数据
    for row_idx, row in enumerate(rows, 2):
        if isinstance(row, dict):
            for col, name in enumerate(fieldnames, 1):
                value = row.get(name)
                if isinstance(value, (list, dict)):
                    value = json.dumps(value, ensure_ascii=False)
                ws.cell(row=row_idx, column=col, value=value)
    
    wb.save(filepath)
    return filepath


def export_sqlite(data: Union[Dict, List[Dict]], 
                  filepath: str,
                  table: str = "data",
                  if_exists: str = "replace",
                  **kwargs) -> str:
    """
    导出数据到 SQLite 数据库
    
    Args:
        data: 要导出的数据（字典或字典列表）
        filepath: 数据库文件路径
        table: 表名
        if_exists: 如果表存在的处理方式 ("replace", "append", "fail")
        
    Returns:
        输出文件的绝对路径
    """
    import sqlite3
    
    filepath = os.path.abspath(filepath)
    
    # 统一转换为列表
    if isinstance(data, dict):
        has_list_values = any(isinstance(v, list) for v in data.values())
        if has_list_values:
            rows = _expand_dict_with_lists(data)
        else:
            rows = [data]
    elif isinstance(data, list):
        rows = data
    else:
        rows = [{"value": data}]
    
    if not rows:
        # 创建空数据库
        conn = sqlite3.connect(filepath)
        conn.close()
        return filepath
    
    # 获取所有字段名
    fieldnames = []
    for row in rows:
        if isinstance(row, dict):
            for key in row.keys():
                if key not in fieldnames:
                    fieldnames.append(key)
    
    conn = sqlite3.connect(filepath)
    cursor = conn.cursor()
    
    # 处理表存在的情况
    if if_exists == "replace":
        cursor.execute(f"DROP TABLE IF EXISTS {table}")
    elif if_exists == "fail":
        cursor.execute(f"SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table,))
        if cursor.fetchone():
            conn.close()
            raise ValueError(f"Table '{table}' already exists")
    
    # 创建表（如果不存在）
    columns = ", ".join([f'"{name}" TEXT' for name in fieldnames])
    cursor.execute(f"CREATE TABLE IF NOT EXISTS {table} ({columns})")
    
    # 插入数据
    placeholders = ", ".join(["?" for _ in fieldnames])
    insert_sql = f"""INSERT INTO {table} ({', '.join([f'"{n}"' for n in fieldnames])}) VALUES ({placeholders})"""
    
    for row in rows:
        if isinstance(row, dict):
            values = []
            for name in fieldnames:
                value = row.get(name)
                if isinstance(value, (list, dict)):
                    value = json.dumps(value, ensure_ascii=False)
                elif value is not None:
                    value = str(value)
                values.append(value)
            cursor.execute(insert_sql, values)
    
    conn.commit()
    conn.close()
    
    return filepath


def save_response(content: Union[str, bytes], 
                  filepath: str,
                  encoding: str = "utf-8") -> str:
    """
    保存响应内容到文件
    
    Args:
        content: 响应内容（字符串或字节）
        filepath: 输出文件路径
        encoding: 文件编码（仅用于字符串）
        
    Returns:
        输出文件的绝对路径
    """
    filepath = os.path.abspath(filepath)
    
    if isinstance(content, bytes):
        with open(filepath, 'wb') as f:
            f.write(content)
    else:
        with open(filepath, 'w', encoding=encoding) as f:
            f.write(content)
    
    return filepath

